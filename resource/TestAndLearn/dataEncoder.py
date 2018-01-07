'''
Created on 4 Jan. 2018

@author: Alvin UTS
'''
import numpy as np
from openpyxl import load_workbook
import datetime
from sklearn import preprocessing

LColumn = 4


wb = load_workbook(filename = 'D://AAA Routine Backup//PHD//Learning Progress//Metric Learning//demographic+Data+From+mimic.xlsx')
ws = wb.active
colNo = len(list(ws.columns))
rowNo = 10

i = 0
for row in ws.iter_rows(min_row=2, max_col = colNo, max_row=rowNo):
    i = i + 1
    curRow =list()
    for cell in row:
        if cell.value is None:
            cell.value = "Missing Value"
        curRow.append(cell.value)
    if i == 1:
        Data = np.array([curRow],dtype=object)
    else:
        Data = np.insert(Data, len(Data), values=curRow, axis=0)

print ("Original Data", "\n", Data)
#print(len(TrainData.T[15]),'\n')

for i in range (0, len(Data)):
    if isinstance(Data.T[8][i], datetime.date):
        Data[8][i] = '0 Day'
    for s in (Data.T[8][i]).split():
        if s.isdigit():
            Data.T[8][i] = int(s)

print ("Data after 0 day", "\n", Data)


le = preprocessing.LabelEncoder()
LabelName = list()
LabelCode = list()
for i in range (0,8):
    le.fit(Data.T[i])
    #print(list(le.classes_))
    LabelName.append(list(le.classes_))
    LabelCode.append(list(le.transform(list(le.classes_))))
    

    Data.T[i] = le.transform(Data.T[i]) 
    #print(TrainData[2])
    #array([2, 2, 1]...)
    #print(list(le.inverse_transform([2, 2, 1])))
#print(TrainData)
#print('sssss2', len(TrainData[0]))
'''
print (LabelName[7])
print (LabelCode[7])
print (len(LabelName[7]))
print (len(LabelCode[7]))
'''
print ("Data after LabelEncoder ", "\n", Data)

#print(TrainData[0:5])

#def column(matrix, i):
    #return [row[i] for row in matrix]
    
#labels = np.array(column(Data, LColumn))
#Data = np.delete(Data, (LColumn), axis = 1)
#print(labels,'\n')
#print(len(labels))
#if LColumn < 8:
    #print(LabelName[LColumn],'\n')

#print('sssss3', len(TrainData[0]))
#print(TrainData)

#print(len(Trainlabels))
#print(len(TrainData[0]))
#print(TrainData[0:8])
'''
OneHotEncoder
'''

from sklearn.preprocessing import OneHotEncoder
if LColumn < 8:
    enc = OneHotEncoder(categorical_features= np.array([0,1,2,3,4,5,6]))
    enc.fit(Data)
    FinalData = enc.transform(Data).toarray()
else:
    enc = OneHotEncoder(categorical_features= np.array([0,1,2,3,4,5,6,7]))
    enc.fit(Data)
    FinalData = enc.transform(Data).toarray()



